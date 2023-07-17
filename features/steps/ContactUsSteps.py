from behave import *
from config.config import TestData
from pages.ContactUsPage import ContactUsPage
import openpyxl


@given(u'user navigates to contact us page')
def step_impl(context):
    context.contact_page = ContactUsPage(context.driver)
    context.driver.get(TestData.URL)
    context.contact_page.click_on_contactus()


@when(u'user fills the form from given sheetname "{subjectHeading}" and rownumber "{rownum}"')
def step_impl(context, subjectHeading, rownum):
    data_file = openpyxl.load_workbook(r"C:\Users\kumbh\PycharmProjects\BehaveSeleniumBDDFramework\resources\Data.xlsx")
    data_file_worksheet = data_file.active
    # subject_heading = data_file_worksheet.cell(row=int(rownum), column=1)
    message = data_file_worksheet.cell(row=int(rownum), column=2)
    context.contactus_page = ContactUsPage(context.driver)
    context.contactus_page.select_subject_Heading(subjectHeading)
    context.contactus_page.enter_message(message.value)


@when(u'user clicks on send button')
def step_impl(context):
    context.contactus_page.click_on_send()


@then(u'it shows a successful message "{expected_message}"')
def step_impl(context, expected_message):
    assert context.contactus_page.get_message() == expected_message

